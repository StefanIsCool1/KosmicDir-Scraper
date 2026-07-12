"""Validation suite for exporter.py — the /discover final deliverable.

Run:  python3 test_final_export.py

Covers:
  - cross-source consolidation: dedupe-merge by name (gaps filled), nameless
    records kept only with contact data, empty records dropped
  - "neat" JSON: canonical key order, empty fields dropped, records sorted,
    per-record source_url, metadata with per-source counts + field coverage
  - Phase2-Dump files found via the search-dir list (the old merge silently
    skipped them)
  - Excel workbook: Members + Summary sheets, headers, row counts
  - CSV behavior preserved (app._records_to_csv alias still importable)
  - /download serves .xlsx binaries and rejects path-shaped filenames
"""

import json
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "Bot"))

os.environ["SCRAPER_HEADLESS"] = "1"

from exporter import (  # noqa: E402
    export_final_dataset, consolidate, tidy_member, records_to_csv, FIELD_ORDER,
)

_FAILURES = []
_PASSES = 0


def check(name: str, cond: bool, detail: str = ""):
    global _PASSES
    if cond:
        _PASSES += 1
        print(f"  PASS  {name}")
    else:
        _FAILURES.append(f"{name}  {detail}")
        print(f"  FAIL  {name}  {detail}")


def _write(dirpath, fname, members, source_url):
    path = os.path.join(dirpath, fname)
    with open(path, "w") as f:
        json.dump({
            "metadata": {"source_url": source_url, "total_members": len(members)},
            "members": members,
        }, f)
    return path


def _fixture_dirs(td):
    """dirA plays Data-dump, dirB plays Phase2-Dump."""
    dir_a = os.path.join(td, "Data-dump")
    dir_b = os.path.join(td, "Phase2-Dump")
    os.makedirs(dir_a)
    os.makedirs(dir_b)

    _write(dir_a, "dira_structured.json", [
        {"company_name": "Acme Decks", "phone": "555-0100", "website": None,
         "description": "", "contacts": []},
        {"company_name": "Zeta Builders", "website": "https://zeta.example.com"},
        # intra-file duplicate of Acme with extra data (messy name spacing)
        {"company_name": "  acme decks. ",
         "contacts": [{"name": None, "email": "info@acmedecks.com"}]},
    ], "https://dir-a.example.com/members")

    _write(dir_b, "dirb_enriched.json", [
        # cross-source duplicate of Acme — fills website/social/description
        {"company_name": "Acme Decks", "website": "https://acmedecks.com",
         "description": "Deck builder in Dallas.",
         "social_media": {"facebook": "https://fb.com/acme", "twitter": ""},
         "enrichment_status": "enriched"},
        {"company_name": "Beta Porches", "phone": "555-0200",
         "street_address": "1 Main St, Austin, TX"},
        # nameless but has a phone → kept as anonymous record
        {"company_name": "", "phone": "555-0300"},
        # nameless AND contactless → dropped
        {"company_name": "", "description": "junk row"},
    ], "https://dir-b.example.com/roster")

    return dir_a, dir_b


def test_consolidate_and_json():
    print("\n[consolidate + final JSON]")
    with tempfile.TemporaryDirectory() as td:
        dir_a, dir_b = _fixture_dirs(td)

        info = export_final_dataset(
            ["dira_structured.json", "dirb_enriched.json"],
            [dir_a, dir_b], dir_a, "deck_contractors_TX",
            goal="deck contractors in TX", industry="deck contractors",
            locations=["TX"],
        )

        check("export returns info", info is not None)
        check("json + xlsx names", info["json_file"] == "deck_contractors_TX_final.json"
              and info["xlsx_file"] == "deck_contractors_TX_final.xlsx", str(info))

        with open(os.path.join(dir_a, info["json_file"])) as f:
            data = json.load(f)
        meta, members = data["metadata"], data["members"]

        check("4 unique records (2 dups merged, 1 junk dropped)",
              len(members) == 4 and meta["duplicates_merged"] == 2
              and meta["records_dropped_empty"] == 1,
              f"n={len(members)} dups={meta['duplicates_merged']} dropped={meta['records_dropped_empty']}")

        names = [m.get("company_name", "") for m in members]
        check("sorted by name, nameless last",
              names[:3] == ["Acme Decks", "Beta Porches", "Zeta Builders"]
              and names[3] == "", str(names))

        acme = members[0]
        check("duplicate gaps filled across sources",
              acme.get("phone") == "555-0100"
              and acme.get("website") == "https://acmedecks.com"
              and acme.get("description") == "Deck builder in Dallas."
              and acme.get("contacts") == [{"name": None, "email": "info@acmedecks.com"}],
              json.dumps(acme))
        check("empty social values dropped",
              acme.get("social_media") == {"facebook": "https://fb.com/acme"},
              str(acme.get("social_media")))
        check("first source wins provenance",
              acme.get("source_url") == "https://dir-a.example.com/members",
              str(acme.get("source_url")))

        canonical_positions = [FIELD_ORDER.index(k) for k in acme.keys()
                               if k in FIELD_ORDER]
        check("keys follow canonical order",
              canonical_positions == sorted(canonical_positions), str(list(acme.keys())))
        check("no empty fields in records",
              all(v not in (None, "", [], {})
                  for m in members for v in m.values()))

        check("metadata per-source counts",
              len(meta["sources"]) == 2
              and meta["sources"][0]["records"] == 3
              and meta["sources"][1]["records"] == 4, str(meta["sources"]))
        check("field coverage computed",
              meta["field_coverage"]["with_phone"] == 3
              and meta["field_coverage"]["with_email"] == 1,
              str(meta["field_coverage"]))
        return os.path.join(dir_a, info["xlsx_file"]), members


def test_xlsx():
    print("\n[Excel workbook]")
    with tempfile.TemporaryDirectory() as td:
        dir_a, dir_b = _fixture_dirs(td)
        info = export_final_dataset(
            ["dira_structured.json", "dirb_enriched.json"],
            [dir_a, dir_b], dir_a, "deck_contractors_TX",
            goal="deck contractors in TX", industry="deck contractors",
            locations=["TX"],
        )
        from openpyxl import load_workbook
        wb = load_workbook(os.path.join(dir_a, info["xlsx_file"]))
        check("Members + Summary sheets", wb.sheetnames == ["Members", "Summary"],
              str(wb.sheetnames))

        ws = wb["Members"]
        headers = [c.value for c in ws[1]]
        check("header starts with Company Name", headers[0] == "Company Name", str(headers))
        check("source column present", "Source Url" in headers, str(headers))
        check("one row per record + header", ws.max_row == 4 + 1, f"rows={ws.max_row}")
        check("frozen header row", ws.freeze_panes == "A2", str(ws.freeze_panes))

        name_col = headers.index("Company Name") + 1
        first_name = ws.cell(row=2, column=name_col).value
        check("rows sorted (Acme first)", first_name == "Acme Decks", str(first_name))

        summary = wb["Summary"]
        labels = [summary.cell(row=r, column=1).value for r in range(1, summary.max_row + 1)]
        check("summary lists goal + sources",
              "Goal" in labels and "Sources" in labels and "Field coverage" in labels,
              str(labels))


def test_csv_compat():
    print("\n[CSV compatibility]")
    members = [
        {"company_name": "Acme", "phone": "555", "website": "a.com",
         "contacts": [{"name": "Jo", "email": "jo@a.com"}]},
        {"company_name": "Beta", "street_address": "1 Main St"},
    ]
    csv_text = records_to_csv(members)
    header = csv_text.splitlines()[0]
    check("business CSV header order",
          header.startswith("company_name,website,phone,contact_name,contact_email"),
          header)
    check("empty columns dropped", "facebook" not in header and "source_url" not in header,
          header)

    from app import _records_to_csv
    check("app._records_to_csv alias intact",
          _records_to_csv(members) == csv_text)


def test_download_endpoint():
    print("\n[/download xlsx + traversal guard]")
    from app import app, DATA_DUMP

    fname = "test_export_tmp_final.xlsx"
    with tempfile.TemporaryDirectory() as td:
        dir_a, dir_b = _fixture_dirs(td)
        export_final_dataset(
            ["dira_structured.json", "dirb_enriched.json"],
            [dir_a, dir_b], dir_a, "test_export_tmp",
        )
        os.makedirs(DATA_DUMP, exist_ok=True)
        src = os.path.join(dir_a, fname)
        dst = os.path.join(DATA_DUMP, fname)
        with open(src, "rb") as fin, open(dst, "wb") as fout:
            fout.write(fin.read())

    try:
        client = app.test_client()
        resp = client.get(f"/download/{fname}")
        check("xlsx served with 200", resp.status_code == 200, str(resp.status_code))
        check("xlsx content type",
              "spreadsheetml" in resp.headers.get("Content-Type", ""),
              resp.headers.get("Content-Type", ""))
        check("xlsx bytes are a zip (PK)", resp.data[:2] == b"PK")

        resp = client.get("/download/..%2Fapp.py")
        check("path-shaped filename rejected", resp.status_code in (400, 404),
              str(resp.status_code))
    finally:
        if os.path.isfile(dst):
            os.remove(dst)


def main():
    test_consolidate_and_json()
    test_xlsx()
    test_csv_compat()
    test_download_endpoint()

    print("\n" + "=" * 50)
    print(f"{_PASSES} passed, {len(_FAILURES)} failed")
    for f in _FAILURES:
        print(f"  FAILED: {f}")
    sys.exit(1 if _FAILURES else 0)


if __name__ == "__main__":
    main()
