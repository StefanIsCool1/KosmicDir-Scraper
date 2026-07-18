"""Final-deliverable exporter — one cleaned JSON + one Excel workbook per run.

Consolidates every output file a /discover run produced (Phase 1 structured
dumps, Phase 2 enriched dumps, NPI pulls, standalone-site enrichments) into:

    Data-dump/{industry}_{states}_final.json
    Data-dump/{industry}_{states}_final.xlsx

The JSON is the "neat" version of the raw dumps:
  - one record per business, duplicates across sources merged field-by-field
    (first source wins, gaps filled from later sources)
  - canonical key order per record: identity → contact → location →
    web/social → provenance; empty fields dropped entirely
  - records sorted by name; every record tagged with its source_url
  - metadata block with per-source counts and field coverage

The Excel workbook has a styled "Members" sheet (frozen header, autofilter,
sized columns) and a "Summary" sheet (run info, sources, coverage).

Also the single home of the record-flattening logic (CSV_COLUMNS,
flatten_record, records_to_csv) that app.py's /download CSV export uses —
one implementation for CSV and XLSX so the two never drift.

Dependencies: stdlib + openpyxl (optional — without it the JSON is still
written and xlsx is skipped with a console note).
"""

import csv
import io
import json
import os
import re
from datetime import datetime, timezone

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover — openpyxl is in requirements.txt
    _HAS_OPENPYXL = False


# Canonical key order for final JSON records. Unknown keys (dynamic schemas,
# future fields) follow in their original order; empty values are dropped.
FIELD_ORDER = [
    "company_name", "category", "description",
    "phone", "fax", "contacts",
    "website", "street_address", "mailing_address",
    "social_media", "hours", "services", "founded", "team",
    "enrichment_status", "enrichment_source", "website_source",
    "source_url",
]

# Column order for CSV/XLSX — most important fields first.
CSV_COLUMNS = [
    "company_name", "category", "website",
    "phone", "fax",
    "contact_name", "contact_email",
    "street_address", "mailing_address",
    "description",
    # Phase 2 enrichment fields
    "facebook", "linkedin", "instagram", "twitter", "youtube", "yelp", "pinterest", "tiktok",
    "hours", "services", "founded",
    "team",
    "source_url",
    "enrichment_status", "enrichment_source", "website_source",
]


# ───────────────────────────────────────────
#  Record flattening (shared by CSV + XLSX)
# ───────────────────────────────────────────

def flatten_record(record: dict, dynamic: bool = False) -> dict:
    """Flatten a nested member record into a single-level dict for tabular
    output. Contacts merge into contact_name/contact_email (semicolon-
    separated), social_media flattens to per-platform keys, lists join."""
    row = {}

    # Simple fields — copy directly
    for key in ("company_name", "category", "website", "phone", "fax",
                "street_address", "mailing_address", "description",
                "hours", "founded", "source_url",
                "enrichment_status", "enrichment_source", "website_source"):
        val = record.get(key)
        row[key] = str(val).strip() if val else ""

    # Contacts — flatten into semicolon-separated name/email
    contacts = record.get("contacts", [])
    if contacts and isinstance(contacts, list):
        names = [c.get("name", "") for c in contacts if c.get("name")]
        emails = [c.get("email", "") for c in contacts if c.get("email")]
        row["contact_name"] = "; ".join(names)
        row["contact_email"] = "; ".join(emails)
    else:
        row["contact_name"] = ""
        row["contact_email"] = ""

    # Social media — flatten nested dict
    social = record.get("social_media", {}) or {}
    for platform in ("facebook", "linkedin", "instagram", "twitter", "youtube",
                     "yelp", "pinterest", "tiktok"):
        row[platform] = social.get(platform, "") or ""

    # Services — join list
    services = record.get("services", [])
    row["services"] = ", ".join(services) if isinstance(services, list) else str(services or "")

    # Team — flatten to "Name (Title); Name (Title)"
    team = record.get("team", [])
    if team and isinstance(team, list):
        parts = []
        for member in team:
            if isinstance(member, dict):
                name = member.get("name", "")
                title = member.get("title", "")
                parts.append(f"{name} ({title})" if title else name)
            elif isinstance(member, str):
                parts.append(member)
        row["team"] = "; ".join(parts)
    else:
        row["team"] = ""

    if dynamic:
        # Pass through AI-chosen fields (e.g. model, chipset, vram, price) that
        # aren't part of the fixed business layout. Nested values are JSON-encoded.
        for k, v in record.items():
            if k in row or k in ("contacts", "social_media", "services", "team"):
                continue
            if isinstance(v, (dict, list)):
                row[k] = json.dumps(v, ensure_ascii=False) if v else ""
            else:
                row[k] = str(v).strip() if v not in (None, "") else ""

    return row


def records_to_rows(records: list, entity_type: str = "business") -> tuple[list, list]:
    """Flatten records and pick the columns that actually carry data.

    Returns (columns, flat_rows). For a non-"business" entity_type, AI-chosen
    dynamic keys with data are appended after the business layout, in
    first-seen order.
    """
    dynamic = entity_type != "business"
    flat_records = [flatten_record(r, dynamic=dynamic) for r in records]

    columns_with_data = []
    for col in CSV_COLUMNS:
        if any(row.get(col) for row in flat_records):
            columns_with_data.append(col)

    if dynamic:
        for row in flat_records:
            for k in row:
                if k not in columns_with_data and any(rr.get(k) for rr in flat_records):
                    columns_with_data.append(k)

    return columns_with_data, flat_records


def records_to_csv(records: list, entity_type: str = "business") -> str:
    """Convert a list of member records to a CSV string (empty columns dropped)."""
    columns, flat_records = records_to_rows(records, entity_type=entity_type)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for row in flat_records:
        writer.writerow(row)
    return output.getvalue()


# ───────────────────────────────────────────
#  Cleaning / merging
# ───────────────────────────────────────────

_WS_RE = re.compile(r"\s+")


def _is_empty(v) -> bool:
    return v is None or v == "" or v == [] or v == {}


def _clean_scalar(v):
    if isinstance(v, str):
        return _WS_RE.sub(" ", v).strip()
    return v


def _norm_name(name: str) -> str:
    """Dedup key for a business name: lowercased, whitespace-collapsed,
    trailing punctuation stripped."""
    return _WS_RE.sub(" ", str(name or "")).strip().strip(".,").lower()


def _dedupe_contacts(contacts: list) -> list:
    seen = set()
    out = []
    for c in contacts or []:
        if not isinstance(c, dict):
            continue
        name = _clean_scalar(c.get("name")) or None
        email = _clean_scalar(c.get("email")) or None
        if not name and not email:
            continue
        key = ((name or "").lower(), (email or "").lower())
        if key in seen:
            continue
        seen.add(key)
        out.append({"name": name, "email": email})
    return out


def _merge_into(base: dict, extra: dict):
    """Fill base's gaps from a duplicate record found in another source.
    Scalars: keep base unless empty. Lists/dicts: union."""
    for k, v in extra.items():
        if _is_empty(v):
            continue
        cur = base.get(k)
        if k == "contacts":
            base[k] = _dedupe_contacts((cur or []) + v)
        elif isinstance(v, dict) and isinstance(cur, dict):
            for dk, dv in v.items():
                if _is_empty(cur.get(dk)) and not _is_empty(dv):
                    cur[dk] = dv
        elif isinstance(v, list) and isinstance(cur, list):
            for item in v:
                if item not in cur:
                    cur.append(item)
        elif _is_empty(cur):
            base[k] = v


def tidy_member(m: dict) -> dict:
    """One clean record: canonical key order, whitespace-normalized scalars,
    deduped contacts, empty fields dropped. Unknown keys keep their order
    after the canonical block (dynamic schemas pass through untouched)."""
    src = dict(m)
    if "contacts" in src:
        src["contacts"] = _dedupe_contacts(src.get("contacts"))
    if isinstance(src.get("social_media"), dict):
        src["social_media"] = {k: v for k, v in src["social_media"].items()
                               if not _is_empty(v)}

    out = {}
    for k in FIELD_ORDER:
        v = _clean_scalar(src.get(k)) if not isinstance(src.get(k), (dict, list)) \
            else src.get(k)
        if not _is_empty(v):
            out[k] = v
    for k, v in src.items():
        if k in FIELD_ORDER:
            continue
        v = _clean_scalar(v) if not isinstance(v, (dict, list)) else v
        if not _is_empty(v):
            out[k] = v
    return out


# ───────────────────────────────────────────
#  Reading run output files
# ───────────────────────────────────────────

def _read_dump(path: str) -> tuple[list, dict]:
    """Read a structured/enriched dump. Handles both shapes:
    {"metadata": ..., "members": [...]} and bare [...]. Local copy of
    Bot/main.py's readers so this module stays stdlib-importable."""
    with open(path, "r") as f:
        data = json.load(f)
    if isinstance(data, list):
        return data, {}
    if isinstance(data, dict):
        return data.get("members") or [], data.get("metadata") or {}
    return [], {}


def _resolve(fname: str, search_dirs: list) -> str | None:
    for d in search_dirs:
        p = os.path.join(d, fname)
        if os.path.isfile(p):
            return p
    return None


def consolidate(paths: list) -> dict:
    """Merge the member lists of several dump files into one clean list.

    Dedup key is the normalized name (name_field for dynamic schemas);
    nameless records survive only if they carry a phone/website/email
    (keyed on that) — otherwise they're junk in a final deliverable.
    """
    entity_type = "business"
    name_field = "company_name"
    merged: dict[str, dict] = {}
    order: list[str] = []
    sources = []
    duplicates_merged = 0
    dropped_empty = 0

    for i, path in enumerate(paths):
        try:
            members, meta = _read_dump(path)
        except Exception:
            continue
        if i == 0 and meta.get("entity_type"):
            entity_type = meta["entity_type"]
            name_field = meta.get("name_field") or name_field
        source_url = meta.get("source_url") or os.path.basename(path)
        source_entry = {
            "source_url": source_url,
            "file": os.path.basename(path),
            "records": len(members),
        }
        if meta.get("partial"):
            # Phase 2 count gate: this source extracted less than the site's
            # stated total — carried through so the deliverable says so.
            source_entry["partial"] = True
            if meta.get("expected_count"):
                source_entry["expected_count"] = meta["expected_count"]
        sources.append(source_entry)

        for m in members:
            if not isinstance(m, dict):
                continue
            record = dict(m)
            record.setdefault("source_url", source_url)

            key = _norm_name(record.get(name_field))
            if not key:
                fallback = (record.get("phone") or record.get("website")
                            or (record.get("contacts") or [{}])[0].get("email"))
                if _is_empty(fallback):
                    dropped_empty += 1
                    continue
                key = f"~anon~{_norm_name(str(fallback))}"

            if key in merged:
                _merge_into(merged[key], record)
                duplicates_merged += 1
            else:
                merged[key] = record
                order.append(key)

    members = [tidy_member(merged[k]) for k in order]
    members.sort(key=lambda m: _norm_name(m.get(name_field)) or "~")

    return {
        "members": members,
        "sources": sources,
        "entity_type": entity_type,
        "name_field": name_field,
        "duplicates_merged": duplicates_merged,
        "dropped_empty": dropped_empty,
        "partial": any(s.get("partial") for s in sources),
    }


# ───────────────────────────────────────────
#  Writers
# ───────────────────────────────────────────

def _field_coverage(members: list) -> dict:
    return {
        "with_name": sum(1 for m in members if m.get("company_name")),
        "with_phone": sum(1 for m in members if m.get("phone")),
        "with_email": sum(1 for m in members
                          if any(c.get("email") for c in m.get("contacts", []))),
        "with_website": sum(1 for m in members if m.get("website")),
        "with_address": sum(1 for m in members
                            if m.get("street_address") or m.get("mailing_address")),
        "with_description": sum(1 for m in members if m.get("description")),
        "with_category": sum(1 for m in members if m.get("category")),
        "with_social": sum(1 for m in members
                           if any((m.get("social_media") or {}).values())),
    }


def write_final_xlsx(path: str, members: list, metadata: dict,
                     entity_type: str = "business") -> str | None:
    """Write the Excel workbook: styled Members sheet + Summary sheet.
    Returns the path, or None when openpyxl is unavailable."""
    if not _HAS_OPENPYXL:
        print("  openpyxl not installed — skipping Excel export "
              "(pip install openpyxl)")
        return None

    columns, rows = records_to_rows(members, entity_type=entity_type)

    wb = Workbook()
    ws = wb.active
    ws.title = "Members"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col in enumerate(columns, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(col, ""))

    # Column widths: fit content, capped so descriptions don't explode the sheet
    for c_idx, col in enumerate(columns, start=1):
        longest = max([len(col)] + [len(str(r.get(col, ""))) for r in rows[:200]])
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(longest + 2, 10), 45)

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    # --- Summary sheet ---
    s = wb.create_sheet("Summary")
    bold = Font(bold=True)

    def add(label, value=""):
        row = s.max_row + 1 if s.max_row > 1 or s.cell(1, 1).value else 1
        s.cell(row=row, column=1, value=label).font = bold
        s.cell(row=row, column=2, value=value)

    add("Generated", metadata.get("generated_at", ""))
    if metadata.get("goal"):
        add("Goal", metadata["goal"])
    if metadata.get("industry"):
        add("Industry", metadata["industry"])
    if metadata.get("locations"):
        add("Locations", ", ".join(str(l) for l in metadata["locations"] if l))
    add("Total records", metadata.get("total_members", len(members)))
    add("Duplicates merged", metadata.get("duplicates_merged", 0))
    add("Sources", len(metadata.get("sources", [])))
    for src in metadata.get("sources", []):
        add(f"  {src.get('source_url', '')[:80]}", f"{src.get('records', 0)} records")
    add("")
    add("Field coverage")
    for field, count in (metadata.get("field_coverage") or {}).items():
        add(f"  {field.replace('_', ' ')}", count)
    s.column_dimensions["A"].width = 42
    s.column_dimensions["B"].width = 60

    wb.save(path)
    return path


def export_final_dataset(output_files: list, search_dirs: list, out_dir: str,
                         base_name: str, goal: str = "", industry: str = "",
                         locations: list | None = None) -> dict | None:
    """Build the run's final deliverable from its output files.

    Writes {base_name}_final.json (always) and {base_name}_final.xlsx (when
    openpyxl is available) into out_dir. Returns an info dict for the SSE
    complete event, or None when no output file could be read.
    """
    paths = []
    for fname in output_files:
        p = _resolve(fname, search_dirs)
        if p:
            paths.append(p)
        else:
            print(f"  Final export: {fname} not found in any dump dir — skipped")
    if not paths:
        return None

    result = consolidate(paths)
    members = result["members"]

    metadata = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "goal": goal,
        "industry": industry,
        "locations": [l for l in (locations or []) if l],
        "total_members": len(members),
        "sources": result["sources"],
        "duplicates_merged": result["duplicates_merged"],
        "records_dropped_empty": result["dropped_empty"],
        "entity_type": result["entity_type"],
        "name_field": result["name_field"],
        "field_coverage": _field_coverage(members),
    }
    if result.get("partial"):
        # Additive (Phase 2 count gate): at least one source extracted less
        # than its site's stated total.
        metadata["partial"] = True

    os.makedirs(out_dir, exist_ok=True)
    json_name = f"{base_name}_final.json"
    with open(os.path.join(out_dir, json_name), "w") as f:
        json.dump({"metadata": metadata, "members": members}, f,
                  indent=2, ensure_ascii=False)

    xlsx_name = None
    try:
        xlsx_path = write_final_xlsx(os.path.join(out_dir, f"{base_name}_final.xlsx"),
                                     members, metadata,
                                     entity_type=result["entity_type"])
        if xlsx_path:
            xlsx_name = os.path.basename(xlsx_path)
    except Exception as e:
        print(f"  Excel export failed (JSON still written): {e}")

    return {
        "json_file": json_name,
        "xlsx_file": xlsx_name,
        "total": len(members),
        "sources": result["sources"],
        "duplicates_merged": result["duplicates_merged"],
        "partial": result.get("partial", False),
    }
